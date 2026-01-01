
import os
import requests
from bs4 import BeautifulSoup # pyright: ignore[reportMissingImports]
import chromadb # pyright: ignore[reportMissingImports]
from chromadb.utils import embedding_functions # pyright: ignore[reportMissingImports]
import time
import re
import json
from urllib.parse import urljoin, urlparse
import logging
import io
import tempfile

# Configuration du logging en premier
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)

# Imports pour embeddings locaux (alternative à OpenAI)
# Utilisation de l'API Hugging Face Inference (gratuite, pas de quota, pas besoin de torch)
SENTENCE_TRANSFORMERS_AVAILABLE = False
_HF_API_AVAILABLE = True  # API Hugging Face toujours disponible

def _get_hf_embedding_function():
    """Crée une fonction d'embedding utilisant l'API Hugging Face (gratuite)"""
    try:
        import requests
        
        class HuggingFaceEmbeddingFunction:
            """Fonction d'embedding utilisant l'API Hugging Face Inference (gratuite)"""
            def __init__(self):
                # Utiliser la nouvelle URL router.huggingface.co
                self.api_url = "https://router.huggingface.co/pipeline/feature-extraction/sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
                self.headers = {"Authorization": f"Bearer {os.environ.get('HF_API_KEY', '')}"} if os.environ.get('HF_API_KEY') else {}
            
            def name(self):
                """Retourne le nom de la fonction d'embedding (requis par ChromaDB)"""
                return "huggingface-api"
            
            def embed_query(self, input):
                """Méthode pour les requêtes (compatible ChromaDB)"""
                return self(input)
            
            def __call__(self, input):
                """Génère les embeddings via l'API Hugging Face avec retry et gestion d'erreurs"""
                if isinstance(input, str):
                    input = [input]
                
                max_retries = 3
                retry_delay = 2  # Délai initial de 2 secondes
                
                for attempt in range(max_retries):
                    try:
                        # Délai entre les requêtes pour éviter les rate limits
                        if attempt > 0:
                            time.sleep(retry_delay * (2 ** (attempt - 1)))  # 2s, 4s, 8s
                        
                        # Utiliser l'API Hugging Face (gratuite, pas de quota)
                        # Timeout de 3 secondes maximum pour traitement rapide
                        response = requests.post(
                            self.api_url,
                            headers=self.headers,
                            json={"inputs": input},
                            timeout=3  # Timeout de 3 secondes pour traitement rapide
                        )
                        
                        if response.status_code == 200:
                            embeddings = response.json()
                            # S'assurer que c'est une liste de listes
                            if isinstance(embeddings, list) and len(embeddings) > 0:
                                if isinstance(embeddings[0], list):
                                    return embeddings
                                else:
                                    return [embeddings]
                            return embeddings
                        elif response.status_code == 503:
                            # Modèle en chargement, attendre et réessayer
                            if attempt < max_retries - 1:
                                wait_time = 10 * (attempt + 1)  # 10s, 20s, 30s
                                logger.info(f"Modèle Hugging Face en chargement (503), attente {wait_time}s avant réessai...")
                                time.sleep(wait_time)
                                continue
                            else:
                                logger.warning(f"Modèle Hugging Face toujours en chargement après {max_retries} tentatives")
                                raise Exception(f"HF API error: 503 (modèle en chargement)")
                        elif response.status_code == 429:
                            # Rate limit, attendre plus longtemps
                            if attempt < max_retries - 1:
                                wait_time = 30 * (attempt + 1)  # 30s, 60s, 90s
                                logger.info(f"Rate limit Hugging Face (429), attente {wait_time}s avant réessai...")
                                time.sleep(wait_time)
                                continue
                            else:
                                logger.warning(f"Rate limit Hugging Face toujours actif après {max_retries} tentatives")
                                raise Exception(f"HF API error: 429 (rate limit)")
                        else:
                            # Autre erreur
                            if attempt < max_retries - 1:
                                logger.warning(f"Erreur API Hugging Face: {response.status_code}, réessai dans {retry_delay * (2 ** attempt)}s...")
                                continue
                            else:
                                logger.warning(f"Erreur API Hugging Face: {response.status_code}")
                                raise Exception(f"HF API error: {response.status_code}")
                    except requests.exceptions.Timeout:
                        if attempt < max_retries - 1:
                            logger.warning(f"Timeout Hugging Face API, réessai dans {retry_delay * (2 ** attempt)}s...")
                            continue
                        else:
                            logger.warning("Timeout Hugging Face API après plusieurs tentatives")
                            raise
                    except Exception as e:
                        if attempt < max_retries - 1:
                            logger.warning(f"Erreur avec Hugging Face API: {e}, réessai dans {retry_delay * (2 ** attempt)}s...")
                            continue
                        else:
                            logger.warning(f"Erreur avec Hugging Face API après {max_retries} tentatives: {e}")
                            raise
                
                # Ne devrait jamais arriver ici, mais au cas où
                raise Exception("Erreur inconnue avec Hugging Face API")
        
        return HuggingFaceEmbeddingFunction()
    except Exception as e:
        logger.warning(f"Impossible de créer la fonction Hugging Face: {e}")
        return None

def _get_ollama_embedding_function():
    """Crée une fonction d'embedding utilisant Ollama (fallback si autres échouent)"""
    try:
        # Déterminer l'URL Ollama selon l'environnement (Local ou Docker)
        if os.path.exists("/app"):
            # Docker
            ollama_url = os.environ.get("OLLAMA_BASE_URL", "http://ollama:11434")
            if "host.docker.internal" in ollama_url:
                ollama_url = "http://ollama:11434"
        else:
            # Local : utiliser localhost
            ollama_url = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
            if "ollama:" in ollama_url or "host.docker.internal" in ollama_url:
                ollama_url = "http://localhost:11434"
        
        class OllamaEmbeddingFunction:
            """Fonction d'embedding utilisant Ollama"""
            def __init__(self):
                self.api_url = f"{ollama_url}/api/embeddings"
                self.model = "llama3.2:1b"
            
            def name(self):
                return "ollama"
            
            def embed_query(self, input):
                return self(input)
            
            def __call__(self, input):
                """Génère les embeddings via Ollama en batch avec timeout de 3 secondes"""
                if isinstance(input, str):
                    input = [input]
                
                # Traiter tous les textes en batch pour optimiser la vitesse
                embeddings = []
                try:
                    # Essayer de traiter en batch (si Ollama le supporte)
                    # Sinon, traiter séquentiellement mais rapidement
                    for text in input:
                        try:
                            response = requests.post(
                                self.api_url,
                                json={"model": self.model, "prompt": text},
                                timeout=3  # Timeout de 3 secondes pour traitement rapide
                            )
                            if response.status_code == 200:
                                result = response.json()
                                embedding = result.get("embedding", [])
                                if embedding and isinstance(embedding, list):
                                    embeddings.append(embedding)
                                else:
                                    embeddings.append([0.0] * 3072)
                            else:
                                embeddings.append([0.0] * 3072)
                        except requests.exceptions.Timeout:
                            # Timeout après 3 secondes, utiliser embedding par défaut
                            embeddings.append([0.0] * 3072)
                        except:
                            embeddings.append([0.0] * 3072)
                except:
                    # En cas d'erreur globale, retourner des embeddings par défaut
                    embeddings = [[0.0] * 3072] * len(input)
                
                return embeddings if len(embeddings) > 1 else embeddings
        
        return OllamaEmbeddingFunction()
    except Exception as e:
        logger.warning(f"Impossible de créer la fonction Ollama embeddings: {e}")
        return None

def _get_sentence_transformer():
    """Import lazy de SentenceTransformer (optionnel, nécessite torch)"""
    global SENTENCE_TRANSFORMERS_AVAILABLE
    try:
        from sentence_transformers import SentenceTransformer
        SENTENCE_TRANSFORMERS_AVAILABLE = True
        logger.info("✅ sentence-transformers disponible (embeddings locaux)")
        return SentenceTransformer
    except (ImportError, Exception) as e:
        logger.debug(f"sentence-transformers non disponible: {e}")
        SENTENCE_TRANSFORMERS_AVAILABLE = False
        return None

# Imports optionnels pour les fichiers
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("PyPDF2 non disponible, extraction PDF désactivée")

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl non disponible, extraction Excel désactivée")

# Import pandas de manière lazy pour éviter les blocages
PANDAS_AVAILABLE = False
_pandas_module = None

def _get_pandas():
    """Import lazy de pandas pour éviter les blocages au démarrage"""
    global PANDAS_AVAILABLE, _pandas_module
    if _pandas_module is None:
        try:
            import pandas as pd
            _pandas_module = pd
            PANDAS_AVAILABLE = True
        except (ImportError, Exception) as e:
            logger.debug(f"Pandas non disponible: {e}")
            PANDAS_AVAILABLE = False
            _pandas_module = False
    return _pandas_module if PANDAS_AVAILABLE else None

try:
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    logger.warning("OCR non disponible, extraction texte depuis images désactivée")

# Sections principales à scraper
MAIN_SECTIONS = [
    "emplois-du-temps",
    "listes-des-groupes",
    "absences-formateurs",
    "calendrier-efm",
    "notes-discipline",
    "résultats-fin-année",
    "stage",
    "parrains-de-groupe",
    "règlement-intérieur",
    "documents",
    "support",
    "scholarvox",
    "activités-para-formation",
    "liens-utiles",
    "contact"
]

BASE_URL = "https://sites.google.com/view/ista-ntic-sm/"
SPECIFIC_URL = "https://sites.google.com/view/ista-ntic-sm/emplois-du-temps?authuser=0"

def _check_url_accessible(url, timeout=5):
    """Vérifie rapidement si une URL est accessible (retourne 200, 301, 302)"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.head(url, timeout=timeout, allow_redirects=True, headers=headers)
        # Accepter 200, 301, 302, 303, 307, 308
        return response.status_code in [200, 301, 302, 303, 307, 308]
    except:
        return False

def discover_pages(base_url, max_depth=3, max_pages=100):
    """Découvre récursivement toutes les pages du site Google Sites
    Ne garde que les pages accessibles (ignore les 404)
    
    Args:
        base_url: URL de base du site
        max_depth: Profondeur maximale de récursion (défaut: 3)
        max_pages: Nombre maximum de pages à découvrir (défaut: 100)
    
    Returns:
        Liste des URLs découvertes et accessibles
    """
    discovered_urls = []  # On vérifiera l'accessibilité avant d'ajouter
    visited = set()
    to_visit = [(base_url, 0)]  # (url, depth)
    valid_count = 0
    invalid_count = 0
    
    # Vérifier d'abord que la page d'accueil est accessible
    if _check_url_accessible(base_url):
        discovered_urls.append(base_url)
        visited.add(base_url)
        valid_count += 1
        logger.info(f"Page d'accueil accessible: {base_url}")
    else:
        logger.warning(f"Page d'accueil non accessible: {base_url}")
        invalid_count += 1
    
    try:
        while to_visit and len(discovered_urls) < max_pages:
            current_url, depth = to_visit.pop(0)
            
            if depth >= max_depth:
                continue
            
            try:
                response = requests.get(current_url, timeout=10, allow_redirects=True)
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Chercher tous les liens internes
                for link in soup.find_all('a', href=True):
                    href = link.get('href', '')
                    if not href:
                        continue
                    
                    # Convertir les liens relatifs en absolus
                    if href.startswith('/'):
                        full_url = urljoin(base_url, href)
                    elif href.startswith('http') and 'sites.google.com/view/ista-ntic-sm' in href:
                        full_url = href
                    elif href.startswith('#'):
                        # Ignorer les ancres
                        continue
                    else:
                        # Essayer de construire l'URL relative
                        try:
                            full_url = urljoin(current_url, href)
                        except:
                            continue
                    
                    # Normaliser l'URL (enlever les fragments et paramètres de tracking)
                    parsed = urlparse(full_url)
                    normalized = f"{parsed.scheme}://{parsed.netloc}{parsed.path}".rstrip('/')
                    
                    # Vérifier que c'est bien une page du site et pas déjà visitée
                    if (normalized not in visited and 
                        normalized.startswith(BASE_URL) and
                        not any(ext in normalized.lower() for ext in ['.pdf', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.zip', '.doc', '.docx'])):
                        
                        visited.add(normalized)  # Marquer comme visitée pour éviter les doublons
                        
                        # Vérifier l'accessibilité avant d'ajouter
                        if _check_url_accessible(normalized):
                            discovered_urls.append(normalized)
                            to_visit.append((normalized, depth + 1))
                            valid_count += 1
                            logger.info(f"Découvert (depth {depth + 1}): {normalized} ✅")
                        else:
                            invalid_count += 1
                            logger.debug(f"Page non accessible (ignorée): {normalized}")
                
                # Pause pour éviter de surcharger le serveur
                time.sleep(0.5)
                
            except requests.exceptions.RequestException as e:
                logger.warning(f"Erreur lors de la découverte de {current_url}: {e}")
                invalid_count += 1
                continue
            except Exception as e:
                logger.warning(f"Erreur inattendue lors de la découverte de {current_url}: {e}")
                invalid_count += 1
                continue
        
        # Vérifier aussi les sections connues (au cas où elles n'ont pas été découvertes)
        for section in MAIN_SECTIONS:
            # Essayer différentes variantes d'URL
            section_variants = [
                urljoin(base_url, section),
                urljoin(base_url, section + '/'),
                urljoin(base_url, section.replace('-', '_')),
            ]
            for section_url in section_variants:
                normalized = section_url.rstrip('/')
                if normalized not in visited and normalized.startswith(BASE_URL):
                    visited.add(normalized)  # Marquer comme visitée
                    # Vérifier l'accessibilité avant d'ajouter
                    if _check_url_accessible(normalized):
                        discovered_urls.append(normalized)
                        valid_count += 1
                        logger.info(f"Ajout section connue: {normalized} ✅")
                    else:
                        invalid_count += 1
                        logger.debug(f"Section non accessible (ignorée): {normalized}")
        
        # Dédupliquer et trier
        unique_urls = list(set(discovered_urls))
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"📊 RÉSUMÉ DE LA DÉCOUVERTE")
        logger.info("=" * 60)
        logger.info(f"✅ Pages accessibles: {valid_count}")
        logger.info(f"❌ Pages non accessibles (ignorées): {invalid_count}")
        logger.info(f"📄 Total de pages à scraper: {len(unique_urls)}")
        logger.info("=" * 60)
        logger.info("")
        return unique_urls
        
    except Exception as e:
        logger.error(f"Erreur lors de la découverte des pages: {e}")
        return discovered_urls if discovered_urls else [base_url]  # Retourner les pages valides trouvées

def extract_pdf_content(pdf_url):
    """Extrait le texte d'un fichier PDF"""
    try:
        response = requests.get(pdf_url, timeout=30, stream=True)
        response.raise_for_status()
        
        # Essayer d'abord avec pdfplumber (meilleur pour les tableaux)
        if PDFPLUMBER_AVAILABLE:
            try:
                with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                    text_parts = []
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            text_parts.append(text)
                        # Extraire aussi les tableaux
                        tables = page.extract_tables()
                        for table in tables:
                            if table:
                                text_parts.append("\n[Tableau]\n" + "\n".join(["\t".join(row) for row in table if row]))
                    return "\n\n".join(text_parts)
            except Exception as e:
                logger.warning(f"pdfplumber échoué pour {pdf_url}, essai PyPDF2: {e}")
        
        # Fallback vers PyPDF2
        if PDF_AVAILABLE:
            pdf_file = io.BytesIO(response.content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text_parts = []
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n\n".join(text_parts)
        
        return None
    except Exception as e:
        logger.warning(f"Erreur lors de l'extraction PDF {pdf_url}: {e}")
        return None

def extract_excel_content(excel_url):
    """Extrait le contenu d'un fichier Excel"""
    try:
        response = requests.get(excel_url, timeout=30, stream=True)
        response.raise_for_status()
        
        pd = _get_pandas()
        if pd is not None:
            excel_file = io.BytesIO(response.content)
            # Lire toutes les feuilles
            excel_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
            text_parts = []
            
            for sheet_name, df in excel_data.items():
                text_parts.append(f"\n[Feuille: {sheet_name}]\n")
                # Convertir le DataFrame en texte formaté
                text_parts.append(df.to_string(index=False))
            
            return "\n".join(text_parts)
        elif EXCEL_AVAILABLE:
            excel_file = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n[Feuille: {sheet_name}]\n")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell for cell in row):
                        text_parts.append("\t".join([str(cell) if cell else "" for cell in row]))
            
            return "\n".join(text_parts)
        
        return None
    except Exception as e:
        logger.warning(f"Erreur lors de l'extraction Excel {excel_url}: {e}")
        return None

def extract_image_text(image_url):
    """Extrait le texte d'une image avec OCR"""
    if not OCR_AVAILABLE:
        return None
    
    try:
        response = requests.get(image_url, timeout=30, stream=True)
        response.raise_for_status()
        
        image = Image.open(io.BytesIO(response.content))
        text = pytesseract.image_to_string(image, lang='fra+eng')
        return text if text.strip() else None
    except Exception as e:
        logger.warning(f"Erreur OCR pour {image_url}: {e}")
        return None

def scrape_page(url, retries=3):
    """Scrape une page et extrait son contenu textuel, y compris les fichiers attachés"""
    for attempt in range(retries):
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            response = requests.get(url, timeout=15, headers=headers, allow_redirects=True)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extraire le titre de la page
            title = soup.find('title')
            title_text = title.get_text(strip=True) if title else "Page sans titre"
            
            # Nettoyer les scripts et styles
            for script in soup(["script", "style", "nav", "header", "footer", "iframe"]):
                script.decompose()
            
            text_parts = []
            
            # Sélecteurs spécifiques pour Google Sites (ordre de priorité)
            content_selectors = [
                # Contenu principal Google Sites
                'div[role="main"]',
                'div[data-attrid]',
                'div.y355Mx',
                'div[data-ved]',
                # Sections de contenu
                'section',
                'article',
                # Divs avec contenu textuel
                'div[class*="content"]',
                'div[class*="text"]',
                'div[class*="page"]',
            ]
            
            # Essayer chaque sélecteur et collecter tout le contenu
            found_content = False
            for selector in content_selectors:
                elements = soup.select(selector)
                if elements:
                    for elem in elements:
                        # Extraire le texte avec préservation de la structure
                        text = elem.get_text(separator='\n', strip=True)
                        if text and len(text.strip()) > 20:
                            # Nettoyer les espaces multiples
                            text = re.sub(r'\s+', ' ', text)
                            text_parts.append(text)
                            found_content = True
                    if found_content:
                        break
            
            # Si aucun contenu trouvé avec les sélecteurs, extraire depuis le body
            if not text_parts:
                main_content = soup.find('main') or soup.find('body')
                if main_content:
                    # Extraire les titres séparément pour préserver la structure
                    headings = main_content.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
                    for heading in headings:
                        heading_text = heading.get_text(strip=True)
                        if heading_text:
                            text_parts.append(f"\n## {heading_text}\n")
                    
                    # Extraire les paragraphes
                    paragraphs = main_content.find_all('p')
                    for para in paragraphs:
                        para_text = para.get_text(strip=True)
                        if para_text and len(para_text) > 10:
                            text_parts.append(para_text)
                    
                    # Extraire les listes
                    lists = main_content.find_all(['ul', 'ol'])
                    for list_elem in lists:
                        items = list_elem.find_all('li')
                        list_text = "\n".join([f"- {item.get_text(strip=True)}" for item in items if item.get_text(strip=True)])
                        if list_text:
                            text_parts.append(list_text)
                    
                    # Si toujours rien, extraire tout le texte
                    if not text_parts:
                        full_text = main_content.get_text(separator='\n', strip=True)
                        if full_text:
                            text_parts.append(full_text)
            
            # Chercher et extraire le contenu des fichiers attachés (PDF, Excel, images)
            file_links = soup.find_all('a', href=True)
            extracted_files = set()  # Pour éviter les doublons
            
            for link in file_links:
                href = link.get('href', '')
                link_text = link.get_text(strip=True)
                
                # Convertir en URL absolue
                if href.startswith('/'):
                    file_url = urljoin(url, href)
                elif href.startswith('http'):
                    file_url = href
                else:
                    continue
                
                # Éviter les doublons
                if file_url in extracted_files:
                    continue
                extracted_files.add(file_url)
                
                # Détecter le type de fichier
                file_url_lower = file_url.lower()
                
                if file_url_lower.endswith('.pdf'):
                    logger.info(f"  → Extraction PDF: {link_text or file_url}")
                    pdf_text = extract_pdf_content(file_url)
                    if pdf_text:
                        text_parts.append(f"\n[Contenu du fichier PDF: {link_text or 'Document PDF'}]\n{pdf_text}")
                
                elif file_url_lower.endswith(('.xlsx', '.xls')):
                    logger.info(f"  → Extraction Excel: {link_text or file_url}")
                    excel_text = extract_excel_content(file_url)
                    if excel_text:
                        text_parts.append(f"\n[Contenu du fichier Excel: {link_text or 'Document Excel'}]\n{excel_text}")
                
                elif file_url_lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
                    if OCR_AVAILABLE:
                        logger.info(f"  → Extraction texte depuis image: {link_text or file_url}")
                        image_text = extract_image_text(file_url)
                        if image_text:
                            text_parts.append(f"\n[Texte extrait de l'image: {link_text or 'Image'}]\n{image_text}")
            
            # Joindre le texte avec des séparateurs clairs
            full_text = "\n\n".join(text_parts)
            
            # Nettoyer le texte final
            full_text = re.sub(r'\n{3,}', '\n\n', full_text)  # Réduire les sauts de ligne multiples
            full_text = re.sub(r' +', ' ', full_text)  # Réduire les espaces multiples
            
            # Détecter la section à partir de l'URL et du titre
            section = "accueil"
            url_lower = url.lower()
            title_lower = title_text.lower()
            
            # Mapping des sections avec variantes
            section_mapping = {
                "emplois du temps": ["emploi", "horaire", "planning", "timetable", "schedule", "emplois-du-temps"],
                "listes des groupes": ["groupe", "liste", "groupes", "listes-des-groupes"],
                "absences formateurs": ["absence", "formateur", "professeur", "enseignant", "absences-formateurs"],
                "calendrier efm": ["calendrier", "efm", "examen", "fin de module", "calendrier-efm"],
                "notes discipline": ["note", "discipline", "sanction", "notes-discipline"],
                "résultats fin année": ["résultat", "résultats", "fin d'année", "bulletin", "résultats-fin-année"],
                "stage": ["stage", "stages", "pfe"],
                "parrains de groupe": ["parrain", "parrains", "parrains-de-groupe"],
                "règlement intérieur": ["règlement", "réglementation", "règles", "règlement-intérieur"],
                "documents": ["document", "documents", "fichier"],
                "support": ["support", "aide", "assistance technique"],
                "scholarvox": ["scholarvox", "bibliothèque", "livre"],
                "activités para-formation": ["activité", "activités", "para-formation", "activités-para-formation"],
                "liens utiles": ["lien", "liens", "ressource", "liens-utiles"],
                "contact": ["contact", "contacter", "adresse", "téléphone", "email"]
            }
            
            # Chercher dans l'URL d'abord
            for sec_name, keywords in section_mapping.items():
                if any(keyword in url_lower for keyword in keywords):
                    section = sec_name
                    break
            
            # Si pas trouvé dans l'URL, chercher dans le titre
            if section == "accueil":
                for sec_name, keywords in section_mapping.items():
                    if any(keyword in title_lower for keyword in keywords):
                        section = sec_name
                        break
            
            return {
                'url': url,
                'title': title_text,
                'content': full_text,
                'section': section
            }
            
        except requests.exceptions.RequestException as e:
            logger.warning(f"Tentative {attempt + 1}/{retries} échouée pour {url}: {e}")
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                logger.error(f"Impossible de scraper {url} après {retries} tentatives")
                return None
        except Exception as e:
            logger.error(f"Erreur inattendue lors du scraping de {url}: {e}")
            return None
    
    return None

def chunk_text_intelligent(text, chunk_size=400, overlap=100):
    """Chunking intelligent par paragraphes sémantiques avec overlap
    OPTIMISÉ pour rapidité : chunks de 300-500 tokens (≈400 caractères)
    
    Args:
        text: Texte à découper
        chunk_size: Taille cible d'un chunk en caractères (défaut: 400 ≈ 300-500 tokens)
        overlap: Nombre de caractères de chevauchement entre chunks (défaut: 100)
    
    Returns:
        Liste de chunks de texte
    """
    if not text or not text.strip():
        return []
    
    # Séparer par paragraphes (double saut de ligne) et sections (titres)
    # D'abord, identifier les sections avec des titres (## ou lignes en majuscules)
    lines = text.split('\n')
    paragraphs = []
    current_para = []
    
    for line in lines:
        line_stripped = line.strip()
        # Détecter les titres (lignes courtes en majuscules ou avec ##)
        is_heading = (line_stripped.startswith('##') or 
                     (len(line_stripped) < 100 and line_stripped.isupper() and len(line_stripped) > 3))
        
        if is_heading and current_para:
            # Finaliser le paragraphe actuel
            paragraphs.append('\n'.join(current_para))
            current_para = [line_stripped]
        elif line_stripped:
            current_para.append(line_stripped)
        elif not line_stripped and current_para:
            # Ligne vide = fin de paragraphe
            paragraphs.append('\n'.join(current_para))
            current_para = []
    
    # Ajouter le dernier paragraphe
    if current_para:
        paragraphs.append('\n'.join(current_para))
    
    # Si pas de paragraphes trouvés, diviser par saut de ligne simple
    if not paragraphs:
        paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
    
    # Si toujours rien, diviser par phrases
    if not paragraphs:
        sentences = re.split(r'[.!?]+\s+', text)
        paragraphs = [s.strip() for s in sentences if s.strip()]
    
    chunks = []
    current_chunk = []
    current_length = 0
    
    for para in paragraphs:
        para_length = len(para)
        
        # Si le paragraphe seul dépasse la taille, le diviser intelligemment
        if para_length > chunk_size:
            # Finaliser le chunk actuel s'il existe
            if current_chunk:
                chunks.append('\n'.join(current_chunk))
                current_chunk = []
                current_length = 0
            
            # Diviser le long paragraphe par phrases ou mots
            # Essayer de diviser par phrases d'abord
            sentences = re.split(r'([.!?]+\s+)', para)
            sentence_parts = []
            for i in range(0, len(sentences), 2):
                if i + 1 < len(sentences):
                    sentence_parts.append(sentences[i] + sentences[i + 1])
                elif i < len(sentences):
                    sentence_parts.append(sentences[i])
            
            temp_chunk = []
            temp_length = 0
            
            for sentence in sentence_parts:
                sent_length = len(sentence)
                
                if temp_length + sent_length <= chunk_size:
                    temp_chunk.append(sentence)
                    temp_length += sent_length + 1
                else:
                    if temp_chunk:
                        chunks.append(' '.join(temp_chunk))
                        # Overlap: garder la fin du chunk précédent
                        overlap_text = ' '.join(temp_chunk[-overlap//50:]) if len(temp_chunk) > 1 else ""
                        temp_chunk = [overlap_text, sentence] if overlap_text else [sentence]
                        temp_length = len(' '.join(temp_chunk))
                    else:
                        # Même une phrase est trop longue, diviser par mots
                        words = sentence.split()
                        for word in words:
                            if temp_length + len(word) + 1 <= chunk_size:
                                temp_chunk.append(word)
                                temp_length += len(word) + 1
                            else:
                                if temp_chunk:
                                    chunks.append(' '.join(temp_chunk))
                                temp_chunk = [word]
                                temp_length = len(word)
            
            if temp_chunk:
                current_chunk = temp_chunk
                current_length = temp_length
        else:
            # Vérifier si on peut ajouter ce paragraphe au chunk actuel
            if current_length + para_length + 1 <= chunk_size:
                current_chunk.append(para)
                current_length += para_length + 1
            else:
                # Finaliser le chunk actuel
                if current_chunk:
                    chunks.append('\n'.join(current_chunk))
                
                # Commencer un nouveau chunk avec overlap
                if chunks and overlap > 0:
                    # Prendre la fin du chunk précédent pour overlap
                    prev_chunk = chunks[-1]
                    # Prendre les derniers caractères pour overlap
                    overlap_text = prev_chunk[-overlap:] if len(prev_chunk) > overlap else prev_chunk
                    # Essayer de commencer à une frontière de phrase
                    overlap_sentences = re.split(r'[.!?]+\s+', overlap_text)
                    if len(overlap_sentences) > 1:
                        overlap_text = ' '.join(overlap_sentences[-2:])
                    
                    current_chunk = [overlap_text, para] if overlap_text.strip() else [para]
                    current_length = len('\n'.join(current_chunk))
                else:
                    current_chunk = [para]
                    current_length = para_length
    
    # Ajouter le dernier chunk
    if current_chunk:
        chunks.append('\n'.join(current_chunk))
    
    # Filtrer et nettoyer les chunks
    cleaned_chunks = []
    for chunk in chunks:
        chunk_cleaned = chunk.strip()
        # Ignorer les chunks trop courts (moins de 50 caractères) sauf s'ils sont importants
        if len(chunk_cleaned) >= 50 or any(keyword in chunk_cleaned.lower() for keyword in ['titre', 'section', 'important', 'note']):
            cleaned_chunks.append(chunk_cleaned)
    
    return cleaned_chunks

def ingest_website(update_only=False, resume_from_backup=False):
    """Ingère toutes les pages du site web"""
    logger.info("Début de l'ingestion du site web...")
    
    try:
        from app.rag.pipeline import get_chromadb_path
        chroma_db_path = get_chromadb_path()
    except:
        if os.path.exists("/app"): chroma_db_path = "/app/chroma_db"
        else:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            root_dir = os.path.dirname(os.path.dirname(os.path.dirname(current_dir)))
            chroma_db_path = os.path.join(root_dir, "chroma_db")
        os.makedirs(chroma_db_path, exist_ok=True)
    
    client = chromadb.PersistentClient(path=chroma_db_path)
    
    # PRIORITÉ: sentence-transformers local (gratuit, rapide, pas de quota)
    embedding_function = None
    SentenceTransformer = _get_sentence_transformer()
    if SentenceTransformer is not None:
        try:
            logger.info("✅ Utilisation de sentence-transformers local")
            embedding_model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
            
            class LocalEmbeddingFunction:
                def name(self): return "sentence-transformers"
                def __call__(self, input):
                    if isinstance(input, str): input = [input]
                    embeddings = embedding_model.encode(input, show_progress_bar=False, convert_to_numpy=True, batch_size=32)
                    return embeddings.tolist()
            
            embedding_function = LocalEmbeddingFunction()
        except Exception as e:
            logger.warning(f"Erreur sentence-transformers: {e}")

    # Fallback 1: Hugging Face API
    if not embedding_function:
        hf_ef = _get_hf_embedding_function()
        if hf_ef:
            try:
                hf_ef(["test"])
                embedding_function = hf_ef
                logger.info("✅ Utilisation de Hugging Face API")
            except: pass

    # Fallback 2: Ollama
    if not embedding_function:
        ollama_ef = _get_ollama_embedding_function()
        if ollama_ef:
            embedding_function = ollama_ef
            logger.info("✅ Utilisation d'Ollama")

    if not embedding_function:
        logger.error("❌ Aucune fonction d'embedding disponible")
        return

    # Création/Récupération de la collection
    collection = client.get_or_create_collection(
        name="website_content",
        embedding_function=embedding_function
    )
        
    # Fichier de sauvegarde
    try:
        from app.rag.pipeline import get_chromadb_path
        chroma_db_path = get_chromadb_path()
    except:
        pass  # chroma_db_path déjà défini plus haut
    backup_file = os.path.join(chroma_db_path, "pending_chunks.json")
    
    # Si on reprend depuis une sauvegarde
    all_chunks = []
    all_metadatas = []
    all_ids = []
    
    if resume_from_backup:
        if os.path.exists(backup_file):
            logger.info(f"Reprise depuis la sauvegarde: {backup_file}")
            try:
                with open(backup_file, 'r', encoding='utf-8') as f:
                    backup_data = json.load(f)
                
                all_chunks = backup_data.get('chunks', [])
                all_metadatas = backup_data.get('metadatas', [])
                all_ids = backup_data.get('ids', [])
                
                logger.info(f"Chargement de {len(all_chunks)} chunks depuis la sauvegarde...")
                # On saute directement à la partie ajout à ChromaDB
            except Exception as e:
                logger.error(f"Erreur lors du chargement de la sauvegarde: {e}")
                resume_from_backup = False
    
    if not resume_from_backup:
        # Découvrir toutes les pages
        urls = discover_pages(BASE_URL)
        
        # Ajouter explicitement l'URL spécifique si non présente
        if SPECIFIC_URL not in urls:
            urls.append(SPECIFIC_URL)
            
        logger.info(f"Pages à scraper: {len(urls)}")
        
        # Si update_only, récupérer les URLs déjà ingérées
        existing_urls = set()
        if update_only:
            try:
                existing_data = collection.get()
                if existing_data and existing_data.get('metadatas'):
                    existing_urls = {meta.get('url', '') for meta in existing_data['metadatas']}
            except Exception as e:
                logger.warning(f"Impossible de récupérer les URLs existantes: {e}")
        
        chunk_id = len(all_chunks)  # Continuer la numérotation si on reprend
        
        # Scraper chaque page
        for url in urls:
            if update_only and url in existing_urls:
                logger.info(f"Saut de {url} (déjà ingérée)")
                continue
                
            logger.info(f"Scraping {url}...")
            page_data = scrape_page(url)
            
            if page_data and page_data['content']:
                # Chunking intelligent
                chunks = chunk_text_intelligent(page_data['content'])
                logger.info(f"  → {len(chunks)} chunks créés")
                
                # Préparer les métadonnées
                for i, chunk in enumerate(chunks):
                    all_chunks.append(chunk)
                    all_metadatas.append({
                        "source": page_data['url'],
                        "url": page_data['url'],
                        "title": page_data['title'],
                        "section": page_data['section'],
                        "chunk_index": i,
                        "total_chunks": len(chunks)
                    })
                    all_ids.append(f"chunk_{chunk_id}")
                    chunk_id += 1
                
                # Pause pour éviter de surcharger le serveur
                time.sleep(1)
            else:
                logger.warning(f"  → Aucun contenu extrait de {url}")
    
    # Sauvegarder les chunks scrapés avant de les ajouter à ChromaDB
    if all_chunks and not resume_from_backup:
        try:
            backup_data = {
                "chunks": all_chunks,
                "metadatas": all_metadatas,
                "ids": all_ids,
                "timestamp": time.time()
            }
            with open(backup_file, 'w', encoding='utf-8') as f:
                json.dump(backup_data, f, ensure_ascii=False, indent=2)
            logger.info(f"💾 {len(all_chunks)} chunks sauvegardés dans {backup_file} avant ingestion")
        except Exception as e:
            logger.warning(f"Impossible de sauvegarder les chunks: {e}")
    
    # Ajouter tous les chunks à ChromaDB avec gestion des erreurs de rate limiting
    if all_chunks:
        logger.info(f"Ajout de {len(all_chunks)} chunks à ChromaDB...")
        
        # Traiter les chunks par batch pour optimiser la vitesse (batch de 10 pour traitement rapide)
        batch_size = 10  # Traiter 10 chunks à la fois pour traitement rapide et efficace
        successful_batches = 0
        failed_batches = 0
        pending_chunks = []
        pending_metadatas = []
        pending_ids = []
        
        for i in range(0, len(all_chunks), batch_size):
            batch_chunks = all_chunks[i:i+batch_size]
            batch_metadatas = all_metadatas[i:i+batch_size]
            batch_ids = all_ids[i:i+batch_size]
            
            # Retry logic optimisé pour traitement rapide (max 3 secondes par batch)
            max_retries = 2  # Réduire les tentatives pour traitement rapide
            retry_delay = 0.5  # Délai initial réduit à 0.5 secondes
            
            success = False
            last_error = None
            batch_start_time = time.time()  # Temps de début pour timeout de 3 secondes
            
            # Variable pour suivre si on a déjà essayé de passer à Ollama
            ollama_fallback_tried = False
            
            for attempt in range(max_retries):
                # Vérifier le timeout de 3 secondes
                elapsed_time = time.time() - batch_start_time
                if elapsed_time >= 3.0:
                    logger.warning(f"  ⏱️ Timeout de 3 secondes atteint pour le batch {i//batch_size + 1}")
                    break
                
                try:
                    collection.add(
                        documents=batch_chunks,
                        metadatas=batch_metadatas,
                        ids=batch_ids
                    )
                    elapsed_time = time.time() - batch_start_time
                    logger.info(f"  ✅ Batch {i//batch_size + 1}/{len(all_chunks)//batch_size + 1} ajouté avec succès ({elapsed_time:.2f}s)")
                    successful_batches += len(batch_chunks)
                    success = True
                    break
                    
                except Exception as e:
                    last_error = e
                    error_msg = str(e).lower()
                    
                    # Si erreur 401/403 avec Hugging Face ou erreur d'embedding, recréer la collection avec Ollama
                    is_hf_error = ("401" in error_msg or "403" in error_msg or "unauthorized" in error_msg or 
                                 "hugging" in error_msg or "hf api" in error_msg or 
                                 ("embedding" in error_msg and ("error" in error_msg or "failed" in error_msg)))
                    
                    if is_hf_error and not ollama_fallback_tried:  # Une seule fois pour éviter les boucles
                        ollama_fallback_tried = True
                        logger.warning("⚠️ Erreur avec l'embedding function actuelle, passage à Ollama...")
                        try:
                            # Supprimer l'ancienne collection
                            try:
                                client.delete_collection(name="website_content")
                            except:
                                pass
                            
                            # Créer avec Ollama
                            ollama_ef = _get_ollama_embedding_function()
                            if ollama_ef:
                                collection = client.get_or_create_collection(
                                    name="website_content",
                                    embedding_function=ollama_ef
                                )
                                logger.info("✅ Collection recréée avec embeddings Ollama")
                                # Réessayer l'ajout immédiatement (sans incrémenter attempt)
                                collection.add(
                                    documents=batch_chunks,
                                    metadatas=batch_metadatas,
                                    ids=batch_ids
                                )
                                elapsed_time = time.time() - batch_start_time
                                logger.info(f"  ✅ Batch {i//batch_size + 1}/{len(all_chunks)//batch_size + 1} ajouté avec succès (Ollama) ({elapsed_time:.2f}s)")
                                successful_batches += len(batch_chunks)
                                success = True
                                break
                            else:
                                logger.error("Impossible de créer la fonction Ollama embeddings")
                        except Exception as e2:
                            logger.error(f"Erreur lors du passage à Ollama: {e2}")
                            # Continuer avec les autres tentatives
                    
                    if "rate" in error_msg or "429" in error_msg or "quota" in error_msg:
                        # Pour traitement rapide, ne pas attendre longtemps
                        elapsed_time = time.time() - batch_start_time
                        remaining_time = 3.0 - elapsed_time
                        if attempt < max_retries - 1 and remaining_time > 0.5:
                            wait_time = min(remaining_time * 0.5, 1.0)  # Attendre max 1 seconde
                            logger.warning(f"  ⏳ Quota insuffisant (tentative {attempt + 1}/{max_retries}), attente {wait_time:.1f}s...")
                            time.sleep(wait_time)
                        else:
                            logger.error(f"  ❌ Échec du batch {i//batch_size + 1} après {max_retries} tentatives (quota/timeout)")
                            logger.info(f"     📦 Les chunks seront sauvegardés pour reprise ultérieure")
                            failed_batches += len(batch_chunks)
                            # Ajouter aux chunks en attente seulement si pas déjà dedans
                            if batch_chunks[0] not in [c[:100] for c in pending_chunks[:10] if c]:  # Vérification simple
                                pending_chunks.extend(batch_chunks)
                                pending_metadatas.extend(batch_metadatas)
                                pending_ids.extend(batch_ids)
                            break
                    elif "collection" in error_msg or "not found" in error_msg:
                        # Erreur de collection, réessayer immédiatement (rapide)
                        elapsed_time = time.time() - batch_start_time
                        if elapsed_time < 2.5:  # Réessayer seulement si on a encore du temps
                            logger.warning(f"  ⚠️ Erreur de collection, réessai immédiat...")
                            time.sleep(0.1)
                        else:
                            break
                    else:
                        # Autre type d'erreur
                        elapsed_time = time.time() - batch_start_time
                        remaining_time = 3.0 - elapsed_time
                        logger.error(f"  ❌ Erreur lors de l'ajout du batch {i//batch_size + 1}: {str(e)[:100]}")
                        if attempt < max_retries - 1 and remaining_time > 0.5:
                            wait_time = min(remaining_time * 0.5, 0.5)  # Attendre max 0.5 seconde
                            logger.info(f"     Réessai dans {wait_time:.1f}s...")
                            time.sleep(wait_time)
                        else:
                            failed_batches += len(batch_chunks)
                            if batch_chunks[0] not in [c[:100] for c in pending_chunks[:10] if c]:
                                pending_chunks.extend(batch_chunks)
                                pending_metadatas.extend(batch_metadatas)
                                pending_ids.extend(batch_ids)
                            break
            
            # Sauvegarder progressivement les chunks en attente (tous les 5 chunks ou à la fin)
            if pending_chunks and (len(pending_chunks) % 5 == 0 or i + batch_size >= len(all_chunks)):
                try:
                    backup_data = {
                        "chunks": pending_chunks,
                        "metadatas": pending_metadatas,
                        "ids": pending_ids,
                        "timestamp": time.time(),
                        "total_failed": len(pending_chunks),
                        "total_successful": successful_batches
                    }
                    with open(backup_file, 'w', encoding='utf-8') as f:
                        json.dump(backup_data, f, ensure_ascii=False, indent=2)
                    logger.info(f"  💾 {len(pending_chunks)} chunks en attente sauvegardés")
                except Exception as e:
                    logger.warning(f"Impossible de sauvegarder les chunks en attente: {e}")
            
            # Pause minimale entre les batches (0.1s pour traitement rapide)
            if i + batch_size < len(all_chunks):
                time.sleep(0.1)
        
        # Statistiques finales
        total_batches = len(all_chunks) // batch_size + (1 if len(all_chunks) % batch_size else 0)
        success_rate = (successful_batches / len(all_chunks) * 100) if all_chunks else 0
        
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"📊 RÉSUMÉ DE L'INGESTION (Traitement optimisé - max 3s par batch)")
        logger.info("=" * 60)
        logger.info(f"✅ Chunks réussis: {successful_batches}/{len(all_chunks)} ({success_rate:.1f}%)")
        if failed_batches > 0:
            logger.info(f"❌ Chunks échoués: {failed_batches}/{len(all_chunks)}")
        logger.info(f"⚡ Batch size: {batch_size} chunks par batch")
        logger.info("=" * 60)
        
        # Si tous les chunks ont été ajoutés, supprimer le fichier de sauvegarde
        if successful_batches == len(all_chunks) and os.path.exists(backup_file):
            try:
                os.remove(backup_file)
                logger.info("✅ Tous les chunks ajoutés avec succès, fichier de sauvegarde supprimé")
                logger.info("🎉 L'ingestion est complète ! Le chat peut maintenant utiliser toutes les données.")
            except Exception as e:
                logger.warning(f"Impossible de supprimer le fichier de sauvegarde: {e}")
        
        # Si des chunks ont échoué, informer l'utilisateur avec des instructions claires
        if failed_batches > 0:
            logger.warning("")
            logger.warning("⚠️ ATTENTION: Certains chunks n'ont pas pu être ajoutés")
            logger.warning(f"   {failed_batches} chunks échoués sur {len(all_chunks)} total")
            logger.info("")
            logger.info(f"💾 Les chunks en attente sont sauvegardés dans: {backup_file}")
            logger.info("")
            logger.info("📋 Pour compléter l'ingestion, choisissez une solution:")
            logger.info("")
            logger.info("   Option 1 (Recommandé): Reprendre depuis la sauvegarde")
            logger.info("   ──────────────────────────────────────────────────────")
            logger.info("   Une fois le quota OpenAI restauré, exécutez:")
            logger.info("   docker-compose exec backend python -c \"from app.rag.ingest import ingest_website; ingest_website(resume_from_backup=True)\"")
            logger.info("")
            logger.info("   Option 2: Installer sentence-transformers (évite les quotas)")
            logger.info("   ──────────────────────────────────────────────────────")
            logger.info("   docker-compose exec backend pip install sentence-transformers")
            logger.info("   Puis réexécutez: docker-compose exec backend python -m app.rag.ingest")
            logger.info("")
            logger.info("   Option 3: Augmenter le quota OpenAI")
            logger.info("   ──────────────────────────────────────────────────────")
            logger.info("   Allez sur https://platform.openai.com/account/billing")
            logger.info("")
            logger.info(f"💡 Note: Le chat fonctionne déjà avec les {successful_batches} chunks ajoutés !")
    else:
        logger.warning("Aucun chunk à ajouter!")

if __name__ == "__main__":
    ingest_website()
